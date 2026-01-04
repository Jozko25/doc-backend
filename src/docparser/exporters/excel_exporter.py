"""Excel exporter for canonical documents."""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core.models import CanonicalDocument
from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Export canonical document to Excel format."""

    # Styles
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SECTION_FONT = Font(bold=True, size=12)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @property
    def format_name(self) -> str:
        return "Excel"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    @property
    def mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def export(self, document: CanonicalDocument) -> bytes:
        """
        Export document to Excel bytes.

        Creates an Excel workbook with:
        - Summary sheet with document info
        - Line items sheet
        - Parties sheet (supplier/customer)

        Args:
            document: CanonicalDocument to export

        Returns:
            Excel file content as bytes
        """
        wb = Workbook()

        # Summary sheet (default sheet)
        self._create_summary_sheet(wb.active, document)
        wb.active.title = "Summary"

        # Line Items sheet
        items_ws = wb.create_sheet("Line Items")
        self._create_line_items_sheet(items_ws, document)

        # Parties sheet
        parties_ws = wb.create_sheet("Parties")
        self._create_parties_sheet(parties_ws, document)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _create_summary_sheet(self, ws, document: CanonicalDocument) -> None:
        """Create the summary sheet."""
        row = 1

        # Title
        ws.cell(row=row, column=1, value="Invoice Summary")
        ws.cell(row=row, column=1).font = Font(bold=True, size=16)
        row += 2

        # Document info section
        ws.cell(row=row, column=1, value="Document Information")
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        info_data = [
            ("Document Type", document.document.type.value.replace("_", " ").title()),
            ("Document Number", document.document.number or "-"),
            ("Issue Date", str(document.document.issue_date) if document.document.issue_date else "-"),
            ("Due Date", str(document.document.due_date) if document.document.due_date else "-"),
            ("Currency", document.document.currency),
        ]

        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Supplier summary
        ws.cell(row=row, column=1, value="Supplier")
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Name").font = Font(bold=True)
        ws.cell(row=row, column=2, value=document.supplier.name or "-")
        row += 1
        ws.cell(row=row, column=1, value="Tax ID").font = Font(bold=True)
        ws.cell(row=row, column=2, value=document.supplier.tax_id or "-")
        row += 2

        # Customer summary
        ws.cell(row=row, column=1, value="Customer")
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Name").font = Font(bold=True)
        ws.cell(row=row, column=2, value=document.customer.name or "-")
        row += 1
        ws.cell(row=row, column=1, value="Tax ID").font = Font(bold=True)
        ws.cell(row=row, column=2, value=document.customer.tax_id or "-")
        row += 2

        # Totals section
        ws.cell(row=row, column=1, value="Totals")
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        totals_data = [
            ("Subtotal", float(document.totals.subtotal)),
            ("Total Tax", float(document.totals.total_tax)),
            ("Grand Total", float(document.totals.total_amount)),
        ]

        if document.totals.amount_due is not None:
            totals_data.append(("Amount Due", float(document.totals.amount_due)))

        for label, value in totals_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = '#,##0.00'
            row += 1

        # Tax breakdown
        if document.totals.tax_breakdown:
            row += 1
            ws.cell(row=row, column=1, value="Tax Breakdown")
            ws.cell(row=row, column=1).font = self.SECTION_FONT
            row += 1

            for tax in document.totals.tax_breakdown:
                ws.cell(row=row, column=1, value=f"Rate {tax.rate}%").font = Font(bold=True)
                ws.cell(row=row, column=2, value=float(tax.tax_amount)).number_format = '#,##0.00'
                row += 1

        # Payment info
        if document.payment.reference or document.payment.terms:
            row += 1
            ws.cell(row=row, column=1, value="Payment Information")
            ws.cell(row=row, column=1).font = self.SECTION_FONT
            row += 1

            if document.payment.terms:
                ws.cell(row=row, column=1, value="Terms").font = Font(bold=True)
                ws.cell(row=row, column=2, value=document.payment.terms)
                row += 1
            if document.payment.reference:
                ws.cell(row=row, column=1, value="Reference").font = Font(bold=True)
                ws.cell(row=row, column=2, value=document.payment.reference)
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _create_line_items_sheet(self, ws, document: CanonicalDocument) -> None:
        """Create the line items sheet."""
        headers = [
            "Line",
            "Description",
            "Quantity",
            "Unit",
            "Unit Price",
            "Tax Rate %",
            "Tax Amount",
            "Line Total",
        ]

        # Header row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, item in enumerate(document.line_items, start=2):
            data = [
                item.line_number,
                item.description or "",
                float(item.quantity),
                item.unit or "",
                float(item.unit_price),
                float(item.tax_rate) if item.tax_rate else "",
                float(item.tax_amount),
                float(item.line_total),
            ]

            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER

                # Number formatting
                if col in (3, 5, 7, 8):  # Quantity, Unit Price, Tax Amount, Line Total
                    cell.number_format = '#,##0.00'

        # Totals row
        total_row = len(document.line_items) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(document.totals.total_tax)).number_format = '#,##0.00'
        ws.cell(row=total_row, column=8, value=float(document.totals.total_amount)).number_format = '#,##0.00'

        # Adjust column widths
        widths = [8, 40, 12, 10, 15, 12, 15, 15]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_parties_sheet(self, ws, document: CanonicalDocument) -> None:
        """Create the parties (supplier/customer) sheet."""
        # Headers
        headers = ["Field", "Supplier", "Customer"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Data
        fields = [
            ("Name", document.supplier.name, document.customer.name),
            ("Tax ID", document.supplier.tax_id, document.customer.tax_id),
            ("Registration Number", document.supplier.registration_number, document.customer.registration_number),
            ("Street", document.supplier.address.street, document.customer.address.street),
            ("City", document.supplier.address.city, document.customer.address.city),
            ("Postal Code", document.supplier.address.postal_code, document.customer.address.postal_code),
            ("Country", document.supplier.address.country, document.customer.address.country),
            ("Email", document.supplier.contact.email, document.customer.contact.email),
            ("Phone", document.supplier.contact.phone, document.customer.contact.phone),
            ("IBAN", document.supplier.bank.iban, document.customer.bank.iban),
            ("BIC", document.supplier.bank.bic, document.customer.bank.bic),
        ]

        for row_idx, (field, supplier_val, customer_val) in enumerate(fields, start=2):
            ws.cell(row=row_idx, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=supplier_val or "-")
            ws.cell(row=row_idx, column=3, value=customer_val or "-")

            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = self.BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
