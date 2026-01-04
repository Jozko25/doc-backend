"""Excel and CSV extractor using openpyxl and pandas."""

import io
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..utils.file_handlers import FileType
from .base import BaseExtractor, ExtractionResult


class ExcelExtractor(BaseExtractor):
    """Extract structured data from Excel and CSV files."""

    SUPPORTED_TYPES = {
        FileType.EXCEL_XLSX,
        FileType.EXCEL_XLS,
        FileType.CSV,
    }

    async def extract(self, content: bytes, filename: str | None = None) -> ExtractionResult:
        """
        Extract structured data from Excel or CSV file.

        Args:
            content: File bytes
            filename: Original filename to determine format

        Returns:
            ExtractionResult with structured data
        """
        file_stream = io.BytesIO(content)
        warnings = []

        try:
            # Detect file type
            is_csv = filename and filename.lower().endswith(".csv")
            is_xls = filename and filename.lower().endswith(".xls")

            if is_csv:
                structured_data = self._extract_csv(file_stream)
                source_type = "csv"
            elif is_xls:
                # Old Excel format - use pandas
                structured_data = self._extract_xls(file_stream)
                source_type = "excel_xls"
            else:
                # Default to xlsx
                structured_data = self._extract_xlsx(file_stream)
                source_type = "excel_xlsx"

            # Also create a text representation for LLM
            text_repr = self._to_text(structured_data)

            return ExtractionResult(
                text=text_repr,
                structured_data=structured_data,
                confidence=1.0,  # Structured data is precise
                warnings=warnings,
                source_type=source_type,
            )

        except Exception as e:
            return ExtractionResult(
                text=None,
                warnings=[f"Excel/CSV extraction error: {str(e)}"],
                source_type="excel_error",
            )

    def _extract_xlsx(self, file_stream: io.BytesIO) -> dict[str, Any]:
        """Extract data from XLSX file using openpyxl."""
        wb = load_workbook(file_stream, read_only=True, data_only=True)
        sheets_data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []

            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    rows.append([self._clean_cell(cell) for cell in row])

            if rows:
                sheets_data[sheet_name] = {
                    "headers": rows[0] if rows else [],
                    "data": rows[1:] if len(rows) > 1 else [],
                    "row_count": len(rows),
                }

        wb.close()
        return {"sheets": sheets_data, "sheet_count": len(sheets_data)}

    def _extract_xls(self, file_stream: io.BytesIO) -> dict[str, Any]:
        """Extract data from XLS file using pandas."""
        excel_file = pd.ExcelFile(file_stream)
        sheets_data = {}

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            df = df.dropna(how="all")  # Remove empty rows

            if not df.empty:
                sheets_data[sheet_name] = {
                    "headers": df.columns.tolist(),
                    "data": df.values.tolist(),
                    "row_count": len(df),
                }

        return {"sheets": sheets_data, "sheet_count": len(sheets_data)}

    def _extract_csv(self, file_stream: io.BytesIO) -> dict[str, Any]:
        """Extract data from CSV file."""
        # Try different encodings
        for encoding in ["utf-8", "latin-1", "cp1252"]:
            try:
                file_stream.seek(0)
                df = pd.read_csv(file_stream, encoding=encoding)
                df = df.dropna(how="all")

                return {
                    "sheets": {
                        "Sheet1": {
                            "headers": df.columns.tolist(),
                            "data": df.values.tolist(),
                            "row_count": len(df),
                        }
                    },
                    "sheet_count": 1,
                }
            except UnicodeDecodeError:
                continue

        raise ValueError("Could not decode CSV with any supported encoding")

    def _clean_cell(self, cell: Any) -> Any:
        """Clean cell value for JSON serialization."""
        if cell is None:
            return None
        if isinstance(cell, (int, float, bool)):
            return cell
        if isinstance(cell, str):
            return cell.strip()
        # Convert other types to string
        return str(cell)

    def _to_text(self, structured_data: dict[str, Any]) -> str:
        """Convert structured data to text representation for LLM."""
        lines = []

        for sheet_name, sheet_data in structured_data.get("sheets", {}).items():
            lines.append(f"=== Sheet: {sheet_name} ===")

            headers = sheet_data.get("headers", [])
            data = sheet_data.get("data", [])

            if headers:
                lines.append("Headers: " + " | ".join(str(h) for h in headers))
                lines.append("-" * 50)

            for row_idx, row in enumerate(data[:100]):  # Limit to 100 rows
                row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                lines.append(f"Row {row_idx + 1}: {row_str}")

            if len(data) > 100:
                lines.append(f"... and {len(data) - 100} more rows")

            lines.append("")

        return "\n".join(lines)

    def supports_file_type(self, file_type: str) -> bool:
        """Check if this extractor supports the given file type."""
        try:
            ft = FileType(file_type)
            return ft in self.SUPPORTED_TYPES
        except ValueError:
            return False
